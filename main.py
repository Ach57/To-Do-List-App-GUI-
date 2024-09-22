'''        ----------------- Libraries -----------------        '''
import customtkinter as ctk
from tkinter import messagebox, ttk
from PIL import Image as PILImage
import webbrowser
import tkinter as tk
import time
import threading
import json
import os
import subprocess
import platform
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

'''        ----------------- Constants -----------------        '''

pythonLogoIco_path = './Images/pythonLogo.ico'
pythonLogoJpg_path = './Images/pythonLogo_transparent_bg.png'
githubLogo_path = './Images/Github.png'
pythonLogoImage=''
githubLogoImage=''
text_style = 'Skia'

try:
    pythonLogoImage = ctk.CTkImage(PILImage.open(pythonLogoJpg_path), size=(100,100))
    githubLogoImage = ctk.CTkImage(PILImage.open(githubLogo_path), size=(40,40))
except Exception:
    messagebox.showinfo('Information for user','Not a serious issue. One of the images wasn\'t found and won\'t be displayed. ')
    
    
class mainWindow(ctk.CTk):
    
    def on_hover(self, event):
        self.github.configure(text_color ='#A91B60')
        self.configure(cursor="hand2")  

    def on_leave(self, event):
        self.github.configure(text_color ='white')
        self.configure(cursor="")  

    def goto_github_on_click(self,event):
        webbrowser.open('https://github.com/Ach57')
    
    def update_time(self):
        while True:
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            self.time_label.configure(text=current_time)
            time.sleep(1)
            
    def load_tasks(self):
        if os.path.exists('task.json'):
            with open('task.json','r') as file:
                tasks = json.load(file)
                for task in tasks:
                    self.task_tree.insert('', 'end', values=task)
    
    def submit_task(self):
        task = self.task_entry.get() #Get task name
        if task: 
            current_time = time.strftime("%H:%M:%S")
            current_date= time.strftime("%Y-%m-%d")
            status = 'Incomplete'
            
            self.task_tree.insert('', "end", values=(current_date, current_time,task,status))
            
            self.task_entry.delete(0,"end")
            
        
    def save_tasks(self) -> json:
        """_summary_
        Looks for the elements of the task_tree and saves them in json file
        Returns:
            json: File named task.json
        """
        tasks =[]
        for row in self.task_tree.get_children():
            tasks.append(self.task_tree.item(row)['values'])
        
        with open('task.json','w') as file:
            json.dump(tasks,file)
    
    def on_closing(self):
        if messagebox.askokcancel("Quit", "Are you sure you want to quit ?"):
            self.save_tasks()
            self.destroy()
    
    def delete_selected_task(self):
        is_delete_box_checked = bool(self.delete_task.get()) #True|False
        #Get the selected item
        is_task_selected = self.task_tree.selection()
        
        if is_task_selected:
            if is_delete_box_checked:
                self.task_tree.delete(is_task_selected)
            else:
                messagebox.showinfo('Info','Enusre to check the delete box before deleting a task')
        else:
            messagebox.showinfo('Info','You have to select a task')     
        
       
    def mark_task_completed(self):
        is_mark_complete_checked = bool(self.mark_completed.get())
        is_task_selected = self.task_tree.selection()
        
        if is_task_selected:
            if is_mark_complete_checked:
                current_values = self.task_tree.item(is_task_selected,'values')
                updated_values = list(current_values)
                updated_values[3] = 'Complete'
                self.task_tree.item(is_task_selected,values=updated_values)
            else:
                messagebox.showinfo('Info ','Ensure to check the mark complete box')
        else:
            messagebox.showinfo('Info','You have to select a task')
    
    def get_report(self):
        all_items = self.task_tree.get_children()
        if(all_items):
            data = []
            for item in all_items:
                item_data = self.task_tree.item(item,'values')
                data.append(item_data)
            df = pd.DataFrame(data, columns=['Date','Time','Task','Status'])
            df.to_excel('Task_Report.xlsx',sheet_name='Report',index=False)
            
            workBook = load_workbook('Task_Report.xlsx')
            worksheet = workBook.active
            header_fill = PatternFill(start_color='ADD8E6',end_color='ADD8E6', fill_type='solid')
            header_font = Font(bold=True,color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            column_width = {
                'A': 9.5,
                'B': 7.5,
                'C': 38.17,
                'D': 22.33
            }
            
            for col, width in column_width.items():
                worksheet.column_dimensions[col].width = width
            
            workBook.save('Task_Report.xlsx')
            
            messagebox.showinfo('Message to user','Your report is generated!')
            
            
            if platform.system() =='Windows':
                os.startfile('Task_Report.xlsx')
            elif platform.system() =='Darwin':
                subprocess.call(['open','Task_Report.xlsx'])
            else:
                subprocess.call(['xdg-open', 'Task_Report.xslx'])
            
        else:
            messagebox.showerror('Error','You don\'t have any tasks!')
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        self.iconbitmap(pythonLogoIco_path)
        self.geometry('770x600')
        self._set_appearance_mode('light' )
        self.title('python Task Manager')
        self.protocol('WM_DELETE_WINDOW',self.on_closing)
        
        #Father Farme
        self.father_frame = ctk.CTkFrame(master=self, fg_color='black', corner_radius=20, bg_color='#EBEBEB')
        self.father_frame.pack(pady=15, padx=15, fill='x')
        
        self.father_frame.grid_columnconfigure(1, weight=1)
        
        self.python= ctk.CTkLabel(master=self.father_frame, text='Python',font=('Srisakdi',16,'bold'), text_color='white' )
        self.python.grid(row=0,column=0, padx=20, pady=5)
        
        self.github = ctk.CTkLabel(master=self.father_frame, text='Contact ',
                                   font=('SrisakdiSkia',16,'bold') ,text_color='white',compound='right')
        if githubLogoImage:
            self.github.configure(image = githubLogoImage)
        self.github.grid(row=0, column=2, padx=20, sticky='e')
        
        self.github.bind("<Enter>", self.on_hover)
        self.github.bind('<Leave>',self.on_leave)
        self.github.bind('<Button-1>', self.goto_github_on_click)
        
        
        #Title Frame (Top)
        self.frame = ctk.CTkFrame(master=self, fg_color="#EBEBEB", bg_color="gray", border_color='black', border_width=2)
        self.frame.pack(pady= 10, padx=10, fill='x')
        
        self.title_label = ctk.CTkLabel(master=self.frame, text='Add Task',
                                        image=pythonLogoImage,compound='left', font=(text_style,20,'bold'), text_color='darkblue')
        self.title_label.pack(side = 'left', padx = 10, pady=10)
        
        #Entry for the task
        self.task_entry = ctk.CTkEntry(master=self.frame, width=200, placeholder_text='Type your task here...',
                                       fg_color='white', text_color='black', font=(text_style, 14, 'bold'))
        self.task_entry.pack(side='left', padx=10, pady = 10)
        # Submit Button
        
        self.submit_task_button = ctk.CTkButton(master=self.frame, text='Submit',font=(text_style, 14, 'bold') , command=self.submit_task)
        self.submit_task_button.pack(side='left', padx=10, pady=10)
        
        #Label for current time
        self.time_label = ctk.CTkLabel(master=self.frame, text='', text_color='#A91B60',font=(text_style, 14, 'bold'))
        self.time_label.pack(side='right', padx=10, pady=(0,80))
        
        # Start the time update in a separate thread
        threading.Thread(target=self.update_time, daemon=True).start()
        
        
        
        #Task list Frame (middle)
        self.list_frame = ctk.CTkFrame(master=self, fg_color="#EBEBEB",border_color='black', border_width=2 )
        self.list_frame.pack(pady=10, padx = 10, fill='both', expand = True)
        
        
        style = ttk.Style()
        style.configure("Treeview.Heading",
                        foreground="white",     # Text color for headers
                        font=(text_style, 16, "bold"))  # Font for headers
        
        style.configure("Treeview", foreground='white', font=(text_style,12))
        # Treeview for displaying tasks
        self.task_tree = ttk.Treeview(master=self.list_frame ,columns=('Date', 'Time', 'Task', 'Status'), show='headings')
        self.task_tree.pack(pady=10, padx=10, fill='both', expand=True)
     
        self.task_tree.heading('Date', text='Date')
        self.task_tree.heading('Time', text='Time')
        self.task_tree.heading('Task', text='Task')
        self.task_tree.heading('Status', text='Status')

        self.task_tree.column('Date', width=100)
        self.task_tree.column('Time', width=100)
        self.task_tree.column('Task', width=300)
        self.task_tree.column('Status', width=100)
        
       
        
        
        #Bottom Action Frame
        self.action_frame = ctk.CTkFrame(master=self,fg_color="#EBEBEB",
                                         bg_color="gray", border_color='black', border_width=2,)
        self.action_frame.pack(pady= 10, padx = 10, fill='x')
        
        self.mark_completed = ctk.CTkCheckBox(master=self.action_frame, text='Mark Complete', text_color='black', font=(text_style,14, 'bold'))
        self.mark_completed.pack(side='left', padx=10, pady=10)
        
        self.delete_task = ctk.CTkCheckBox(master=self.action_frame, text='Delete Task', text_color='black', font=(text_style,14,'bold'))
        self.delete_task.pack(side='left', padx=10, pady=10)
        
        
        #Add task button:
        self.mark_task_button = ctk.CTkButton(master=self.action_frame,
                                              text='Mark Task Completed', width=100, font=(text_style,14,'bold'), command=self.mark_task_completed)
        self.mark_task_button.pack(side='left', padx=10, pady=10)
        
        #Delete Task Button 
        self.delete_task_button = ctk.CTkButton(master=self.action_frame, text='Delete Task',
                                                width=100, font=(text_style,14,'bold'), command=self.delete_selected_task)
        self.delete_task_button.pack( side='left',padx=10, pady=10)
        
        #Export report button
        self.export_report_button =ctk.CTkButton(master=self.action_frame, text='Get Report',
                                                 width=100, font=(text_style,14,'bold'), command=self.get_report)
        self.export_report_button.pack(side='left', padx=10, pady=10)
        self.load_tasks()
        
        
    def _on_closing_(self):
        return
    
    def runApp(self) ->ctk.CTk.mainloop:
        self.mainloop()



if __name__=="__main__":
    app = mainWindow()
    app.runApp()
